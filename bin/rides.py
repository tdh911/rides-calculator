import openpyxl 
from people import Person
from car import Car

def format_departure_time(time_in_float):
    is_am = time_in_float < 12
    if is_am:
        hour = int(time_in_float)
        minute = int(float(time_in_float - int(time_in_float)) * 60)
        if minute < 10:
            return "%d:0%d AM" % (hour, minute)
        else:
            return "%d:%d AM" % (hour, minute)
    else:
        hour = int(time_in_float - 12)
        minute = int(float(time_in_float - int(time_in_float)) * 60)
        if minute < 10:
            return "%d:0%d PM" % (hour, minute)
        else:
            return "%d:%d PM" % (hour, minute)


def write_to_excel(cars, unseated_passengers):
    book = openpyxl.Workbook()
    sheet = book.create_sheet(title="Rides List")
    row = 1
    for curr_car in cars:
        driver_name = curr_car.driver.first + " " + curr_car.driver.last
        departure_time_in_float = curr_car.departure_time
        departure_time = format_departure_time(departure_time_in_float)
        capacity = curr_car.seats
        sheet.cell(column=1, row=row, value=driver_name)
        sheet.cell(column=2, row=row, value=departure_time)
        sheet.cell(column=3, row=row, value=capacity)
        row += 1
        for passenger in curr_car.passengers:
            sheet.cell(column=1, row=row, value=passenger.first) 
            sheet.cell(column=2, row=row, value=passenger.last)
            sheet.cell(column=3, row=row, value=passenger.area)
            sheet.cell(column=4, row=row, value=passenger.phone)
            sheet.cell(column=5, row=row, value=passenger.email)
            sheet.cell(column=6, row=row, value=passenger.raw_time)
            row += 1
        row += 1

    sheet.cell(column=1, row=row, value="Unseated passengers")
    row += 1
    for passenger in unseated_passengers:
        sheet.cell(column=1, row=row, value=passenger.first) 
        sheet.cell(column=2, row=row, value=passenger.last)
        sheet.cell(column=3, row=row, value=passenger.area)
        sheet.cell(column=4, row=row, value=passenger.phone)
        sheet.cell(column=5, row=row, value=passenger.email)
        sheet.cell(column=6, row=row, value=passenger.raw_time)
        row += 1

    book.save("RidesList.xlsx")

def main():
    excel_book = raw_input("Type out the excel worksheet filename: ")
    attendees = openpyxl.load_workbook(filename=excel_book, data_only=True)["reg list"]

    max_row = attendees.max_row
    max_column = attendees.max_column

    # Get headers for flexibility
    headers = {}
    for idx in range(1, max_column):
        val = attendees.cell(column=idx, row=1).value
        headers[val] = idx 

    # Construct unsorted list of cars and general passengers
    cars = []
    passengers = []
    for attendee in range(2, max_row):
        if attendees.cell(column=1, row=attendee).value == None:
            break

        first = attendees.cell(column=headers["first"], row=attendee).value.encode("unicode-escape")
        last = attendees.cell(column=headers["last"], row=attendee).value.encode("unicode-escape")
        area = str(attendees.cell(column=headers["area"], row=attendee).value)
        email = str(attendees.cell(column=headers["email"], row=attendee).value)
        phone = str(attendees.cell(column=headers["phone"], row=attendee).value)
        departure_time= str(attendees.cell(column=headers["earliest departure time"], row=attendee).value)

        passenger = Person(first, last, area, phone, email, departure_time)
        seatbelts = int(attendees.cell(column=headers["car seatbelts"], row=attendee).value)
        if not seatbelts == 0:
            car = Car(passenger, seatbelts)
            cars.append(car)
        else:
            passengers.append(passenger)
            
    # Sort both lists based off departure time
    cars = sorted(cars, key=lambda x: x.departure_time, reverse=True)
    passengers = sorted(passengers, key=lambda x: x.earliest_departure, reverse=True)

    car_idx = 0
    # Iterate through the lists and add the passengers to the car
    while len(passengers) > 0:
        curr_car = cars[car_idx]
        passenger = passengers[0]
        if curr_car.add_person(passenger):
            passengers.pop(0)
        else:
            car_idx += 1

        if car_idx == len(cars):
            break

    # Write the cars to an excel, special section for unseated passengers
    write_to_excel(cars, passengers)

if __name__ == "__main__":
    main()
